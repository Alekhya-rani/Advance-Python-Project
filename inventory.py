import tkinter
from PIL import Image, ImageTk
from tkinter import messagebox, ttk
import openpyxl
from openpyxl import Workbook

# GUI Setup
root = tkinter.Tk()
root.title("Inventory Management System")
root.geometry("900x700")
#root.config(bg="#FAD6A5")

bg_image_path = r"C:\Users\alekh\Downloads\Image.jpg"  
bg_image = Image.open(bg_image_path)
bg_image = bg_image.resize((1500,1500), Image.LANCZOS) 
bg_photo = ImageTk.PhotoImage(bg_image)
# Create and place background label
bg_label = tkinter.Label(root, image=bg_photo)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

title_label = tkinter.Label(root, text="Inventory Management System", font=("Arial", 25, "bold"))
title_label.grid(row=0, column=0, columnspan=3, padx=10, pady=10)
title_label.config(anchor="center")  

# Paths for input and output Excel files
input_file_path = r"D:\Alekhya\Inventory_Input.xlsx"
output_file_path = r"D:\Alekhya\Inventory_Output.xlsx"
sheet_name = "InventoryData"

# Load initial data from input file and set up output file
try:
    input_file = openpyxl.load_workbook(input_file_path)
    if sheet_name in input_file.sheetnames:
        input_sheet = input_file[sheet_name]
    else:
        raise ValueError(f"Sheet '{sheet_name}' not found in '{input_file_path}'")
except FileNotFoundError:
    messagebox.showerror("File Error", f"Input file '{input_file_path}' not found.")
    root.destroy()

# Initialize output workbook and copy initial data
output_file = Workbook()
output_sheet = output_file.active
output_sheet.title = sheet_name
output_sheet.append(["ID", "Name", "Quantity", "Price", "Date"])  # Header row

# Load data into Treeview and output sheet from the input file
def load_initial_data():
    for row in input_sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        tree.insert("", "end", values=row)
        output_sheet.append(row)
    output_file.save(output_file_path)  # Save initial data to output file

# Refresh Treeview display from output file
def refresh_table():
    for item in tree.get_children():
        tree.delete(item)
    for row in output_sheet.iter_rows(min_row=2, values_only=True):  # Reload from output
        tree.insert("", "end", values=row)

# Add a new item to the output file and refresh Treeview
def add_item():
    id = id_textbox.get()
    name = name_textbox.get()
    quantity = quantity_textbox.get()
    price = price_textbox.get()
    date = date_textbox.get()

    if not (id and name and quantity and price and date):
        messagebox.showwarning("Input Error", "All fields must be filled!")
        return

    # Add to output sheet and save
    output_sheet.append([id, name, quantity, price, date])
    output_file.save(output_file_path)
    clear_inputs()
    refresh_table()
    messagebox.showinfo("Success", "Item added successfully!")

# Update an existing item in the output file and refresh Treeview
def update_item():
    id = id_textbox.get()
    name = name_textbox.get()
    quantity = quantity_textbox.get()
    price = price_textbox.get()
    date = date_textbox.get()

    if not id:
        messagebox.showwarning("Input Error", "Please enter an ID to update!")
        return

    found = False
    for row in output_sheet.iter_rows(min_row=2):  # Skip header row
        if row[0].value == id:
            # Update only the fields that have been entered by the user
            if name:
                row[1].value = name
            if quantity:
                row[2].value = quantity
            if price:
                row[3].value = price
            if date:
                row[4].value = date
            found = True
            break

    if found:
        output_file.save(output_file_path)
        clear_inputs()
        refresh_table()
        messagebox.showinfo("Success", "Item updated successfully!")
    else:
        messagebox.showwarning("Error", "Item ID not found!")


# Delete an item from the output file and refresh Treeview
def delete_item():
    id = id_textbox.get()

    if not id:
        messagebox.showwarning("Input Error", "Please enter an ID to delete!")
        return

    found = False
    for row in output_sheet.iter_rows(min_row=2):
        if row[0].value == id:
            output_sheet.delete_rows(row[0].row, 1)  # Delete row in output sheet
            found = True
            break

    if found:
        output_file.save(output_file_path)
        clear_inputs()
        refresh_table()
        messagebox.showinfo("Success", "Item deleted successfully!")
    else:
        messagebox.showwarning("Error", "Item ID not found!")

# Retrieve and display item details based on ID
def retrieve_item():
    id = id_textbox.get()

    if not id:
        messagebox.showwarning("Input Error", "Please enter an ID to retrieve!")
        return

    found = False
    for row in output_sheet.iter_rows(min_row=2):
        if row[0].value == id:
            name_textbox.delete(0, tkinter.END)
            name_textbox.insert(0, row[1].value)
            quantity_textbox.delete(0, tkinter.END)
            quantity_textbox.insert(0, row[2].value)
            price_textbox.delete(0, tkinter.END)
            price_textbox.insert(0, row[3].value)
            date_textbox.delete(0, tkinter.END)
            date_textbox.insert(0, row[4].value)
            found = True
            break

    if not found:
        messagebox.showwarning("Error", "Item ID not found!")

# Clear input fields
def clear_inputs():
    id_textbox.delete(0, tkinter.END)
    name_textbox.delete(0, tkinter.END)
    quantity_textbox.delete(0, tkinter.END)
    price_textbox.delete(0, tkinter.END)
    date_textbox.delete(0, tkinter.END)

# Entry fields
tkinter.Label(root, text="ID",font=("Arial",15,"bold")).grid(row=1, column=0, padx=10, pady=10,sticky="e")
id_textbox = tkinter.Entry(root)
id_textbox.grid(row=1, column=1,sticky="w")

tkinter.Label(root, text="Name",font=("Arial",15,"bold")).grid(row=2, column=0, padx=10, pady=10,sticky="e")
name_textbox = tkinter.Entry(root)
name_textbox.grid(row=2, column=1,sticky="w")

tkinter.Label(root, text="Quantity",font=("Arial",15,"bold")).grid(row=3, column=0, padx=10, pady=10,sticky="e")
quantity_textbox = tkinter.Entry(root)
quantity_textbox.grid(row=3, column=1,sticky="w")

tkinter.Label(root, text="Price",font=("Arial",15,"bold")).grid(row=4, column=0, padx=10, pady=10,sticky="e")
price_textbox = tkinter.Entry(root)
price_textbox.grid(row=4, column=1,sticky="w")

tkinter.Label(root, text="Date",font=("Arial",15,"bold")).grid(row=5, column=0, padx=10, pady=10,sticky="e")
date_textbox = tkinter.Entry(root)
date_textbox.grid(row=5, column=1,sticky="w")

# Buttons
tkinter.Button(root, text="Add Item", command=add_item , bg="blue" , fg="white",font=("Arial",10,"bold")).grid(row=6, column=0, padx=10, pady=10)
tkinter.Button(root, text="Update Item", command=update_item , bg="grey" , fg="white",font=("Arial",10,"bold")).grid(row=6, column=1, padx=10, pady=10)
tkinter.Button(root, text="Delete Item", command=delete_item , bg="red" , fg="white",font=("Arial",10,"bold")).grid(row=7, column=1, padx=10, pady=10)
tkinter.Button(root, text="Retrieve Item", command=retrieve_item , bg="green" , fg="white",font=("Arial",10,"bold")).grid(row=7, column=0, padx=10, pady=10)

# Treeview (Table) to display items
tree = ttk.Treeview(root, columns=("ID", "Name", "Quantity", "Price", "Date"), show="headings")
tree.heading("ID", text="ID")
tree.heading("Name", text="Name")
tree.heading("Quantity", text="Quantity")
tree.heading("Price", text="Price")
tree.heading("Date", text="Date")
tree.grid(row=8, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

# Populate the table with initial data from the input file
load_initial_data()

# Run the GUI main loopc
root.mainloop()
